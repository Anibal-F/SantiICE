"""
Generador de Reportes Multi-Cliente CORREGIDO - VERSIÓN LIMPIA CON ERROR FIXED
"""

import pandas as pd
import numpy as np
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Union
import logging
from datetime import datetime
import openpyxl
import openpyxl.utils
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, NamedStyle
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference, PieChart
from openpyxl.chart.label import DataLabelList

# Import config con manejo de errores
try:
    from config import config
except ImportError:
    import sys
    import os
    sys.path.append(os.path.dirname(os.path.abspath(__file__)))
    from config import config

logger = logging.getLogger(__name__)

class ReportGenerator:
    """Generador de reportes personalizado para diferentes tipos de cliente - VERSIÓN CORREGIDA"""
    
    def __init__(self, client_type: str = 'OXXO'):
        """Inicializa el generador de reportes"""
        self.client_type = client_type.upper()
        self.output_dir = Path(config.get_output_dir() if hasattr(config, 'get_output_dir') else './data/output')
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
        logger.info(f"🔧 ReportGenerator INICIALIZANDO con client_type: '{self.client_type}'")
        
        # Configuraciones específicas por cliente
        self._setup_client_configuration()
        
        logger.info(f"📋 Configuración cargada: {self.config['display_name']} - {self.config['report_title']}")
        
        # Configurar estilos de Excel
        self._setup_excel_styles()
        
        logger.info(f"📊 ReportGenerator inicializado para cliente: {self.client_type}")
    
    def _setup_client_configuration(self):
        """Configura parámetros específicos del cliente para reportes"""
        
        logger.info(f"🔧 Configurando parámetros para cliente: {self.client_type}")
        
        client_configs = {
            'OXXO': {
                'display_name': 'OXXO',
                'identifier_name': 'Ticket',
                'identifier_plural': 'Tickets',
                'identifier_field': 'Num. Pedido Adicional',
                'value_field_source': 'Importe Arch. OXXO',
                'value_field_looker': 'Importe Registrado',
                'currency_symbol': '$',
                'report_title': 'REPORTE DE CONCILIACIÓN OXXO',
                'stats_title': 'ESTADÍSTICAS OXXO',
                'missing_source_label': 'No encontrados en OXXO',
                'categories': {
                    'EXACT_MATCH': 'Conciliado',
                    'WITHIN_TOLERANCE': 'Tolerancia',
                    'MINOR_DIFFERENCE': 'Diferencia',
                    'MAJOR_DIFFERENCE': 'Diferencia',
                    'MISSING_IN_OXXO': 'Faltante',
                    'MISSING_IN_LOOKER': 'Faltante'
                },
                'colors': {
                    'primary': '366092',
                    'success': '00AA00', 
                    'warning': 'FFA500',
                    'error': 'FF0000',
                    'info': '0066CC'
                }
            },
            'KIOSKO': {
                'display_name': 'KIOSKO',
                'identifier_name': 'Ticket',
                'identifier_plural': 'Tickets',
                'identifier_field': 'últimos 4 dígitos de ticket',
                'value_field_source': 'Importe KIOSKO',
                'value_field_looker': 'Importe Registrado',
                'currency_symbol': '$',
                'report_title': 'REPORTE DE CONCILIACIÓN KIOSKO',
                'stats_title': 'ESTADÍSTICAS KIOSKO',
                'missing_source_label': 'No encontrados en KIOSKO',
                'categories': {
                    'EXACT_MATCH': 'Conciliado',
                    'WITHIN_TOLERANCE': 'Tolerancia',
                    'MINOR_DIFFERENCE': 'Diferencia',
                    'MAJOR_DIFFERENCE': 'Diferencia',
                    'MISSING_IN_KIOSKO': 'Faltante',
                    'MISSING_IN_LOOKER': 'Faltante',
                    'RETURN_INFORMATIVE': 'Devolución (Informativo)'
                },
                'colors': {
                    'primary': '4A90E2',
                    'success': '7ED321',
                    'warning': 'F5A623',
                    'error': 'D0021B',
                    'info': '50E3C2'
                }
            }
        }
        
        # Asegurar que se use la configuración específica
        if self.client_type in client_configs:
            self.config = client_configs[self.client_type]
            logger.info(f"✅ Configuración específica cargada para {self.client_type}")
            logger.info(f"   📋 Display name: {self.config['display_name']}")
            logger.info(f"   📋 Report title: {self.config['report_title']}")
        else:
            logger.error(f"❌ CLIENTE {self.client_type} NO ENCONTRADO en configuraciones")
            logger.error(f"   Disponibles: {list(client_configs.keys())}")
            raise ValueError(f"Cliente '{self.client_type}' no tiene configuración definida")
    
    def _setup_excel_styles(self):
        """Configura estilos reutilizables para Excel"""
        
        self.excel_styles = {
            'header': {
                'font': Font(bold=True, color="FFFFFF", size=12),
                'fill': PatternFill(start_color=self.config['colors']['primary'], 
                                  end_color=self.config['colors']['primary'], 
                                  fill_type="solid"),
                'alignment': Alignment(horizontal='center', vertical='center'),
                'border': Border()
            },
            'title': {
                'font': Font(bold=True, size=16, color=self.config['colors']['primary']),
                'alignment': Alignment(horizontal='center', vertical='center'),
                'border': Border()
            },
            'subtitle': {
                'font': Font(bold=True, size=14),
                'alignment': Alignment(horizontal='left', vertical='center'),
                'border': Border()
            },
            'currency': {
                'number_format': f'"{self.config["currency_symbol"]}"#,##0.00',
                'border': Border()
            },
            'percentage': {
                'number_format': '0.00%',
                'border': Border()
            },
            'integer': {
                'number_format': '0',
                'border': Border()
            }
        }
    
    def generate_complete_report(self, reconciliation_results: pd.DataFrame, 
                               summary_stats: Dict, timestamp: str = None) -> str:
        """Genera un reporte completo personalizado por cliente usando pandas"""
        
        if timestamp is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        filename = f"reporte_{self.client_type.lower()}_{timestamp}.xlsx"
        filepath = self.output_dir / filename
        
        # Asegurar que el directorio existe
        filepath.parent.mkdir(parents=True, exist_ok=True)
        
        logger.info(f"📊 Generando reporte completo {self.client_type}: {filepath}")
        
        try:
            # Mapear estatus a nombres legibles
            status_mapping = {
                'EXACT_MATCH': 'Conciliado',
                'WITHIN_TOLERANCE': 'Tolerancia', 
                'MINOR_DIFFERENCE': 'Diferencia',
                'MAJOR_DIFFERENCE': 'Diferencia',
                'MISSING_IN_OXXO': 'Faltante',
                'MISSING_IN_KIOSKO': 'Faltante',
                'MISSING_IN_LOOKER': 'Faltante',
                'MISSING_IN_SOURCE': 'Faltante'
            }
            
            # Preparar datos para el reporte
            df_report = reconciliation_results.copy()
            if 'status' in df_report.columns:
                df_report['Estado'] = df_report['status'].map(status_mapping).fillna(df_report['status'])
            
            # Crear resumen ejecutivo
            summary_data = {
                'Métrica': [
                    'Tickets Totales',
                    'Tickets Conciliados', 
                    'Tickets con Diferencia',
                    'Tickets Faltantes',
                    'Tasa de Conciliación (%)',
                    f'Total {self.client_type} ($)',
                    'Total Looker ($)',
                    'Diferencia Total ($)'
                ],
                'Valor': [
                    summary_stats.get('total_records', 0),
                    summary_stats.get('exact_matches', 0),
                    summary_stats.get('major_differences', 0) + summary_stats.get('minor_differences', 0),
                    summary_stats.get('missing_records', 0),
                    f"{summary_stats.get('reconciliation_rate', 0):.1f}%",
                    f"${summary_stats.get('total_client_amount', 0):,.2f}",
                    f"${summary_stats.get('total_looker_amount', 0):,.2f}",
                    f"${summary_stats.get('total_difference', 0):,.2f}"
                ]
            }
            
            df_summary = pd.DataFrame(summary_data)
            
            # Usar pandas ExcelWriter para crear el archivo
            with pd.ExcelWriter(str(filepath), engine='openpyxl') as writer:
                # Hoja de resumen
                df_summary.to_excel(writer, sheet_name='Resumen Ejecutivo', index=False)
                
                # Hoja de resultados detallados
                df_report.to_excel(writer, sheet_name='Resultados Detallados', index=False)
                
                # Hoja de diferencias (solo registros con problemas)
                if 'status' in df_report.columns:
                    df_differences = df_report[~df_report['status'].isin(['EXACT_MATCH'])]
                    if not df_differences.empty:
                        df_differences.to_excel(writer, sheet_name='Diferencias', index=False)
                
                logger.info(f"✅ Reporte Excel generado exitosamente")
            
            # Verificar que el archivo se creó correctamente
            if filepath.exists() and filepath.stat().st_size > 0:
                logger.info(f"📄 Archivo verificado - Tamaño: {filepath.stat().st_size} bytes")
                return str(filepath)
            else:
                raise Exception("El archivo no se generó correctamente")
            
        except Exception as e:
            logger.error(f"❌ Error generando reporte: {e}")
            # Fallback: generar CSV si Excel falla
            try:
                csv_path = filepath.with_suffix('.csv')
                reconciliation_results.to_csv(csv_path, index=False)
                logger.info(f"✅ Archivo CSV generado como fallback: {csv_path}")
                return str(csv_path)
            except Exception as csv_error:
                logger.error(f"❌ Error generando CSV fallback: {csv_error}")
                raise e
    
    def _create_executive_summary_sheet(self, wb: openpyxl.Workbook, 
                                      summary_stats: Dict, timestamp: str):
        """Crea la hoja de resumen ejecutivo CON ORDEN CORREGIDO y SIN LÍNEAS"""
        
        ws = wb.create_sheet("Resumen Ejecutivo")
        ws.sheet_view.showGridLines = False
        
        # Título principal
        ws['A1'] = self.config['report_title']
        ws['A1'].font = Font(size=18, bold=True, color=self.config['colors']['primary'])
        ws['A1'].alignment = Alignment(horizontal='center')
        ws['A1'].border = Border()
        ws.merge_cells('A1:F1')
        
        # Información del reporte
        ws['A3'] = f"Cliente: {self.config['display_name']}"
        ws['A3'].font = Font(size=12, bold=True)
        ws['A3'].border = Border()
        
        ws['A4'] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
        ws['A4'].font = Font(size=11, italic=True)
        ws['A4'].border = Border()
        
        # ORDEN CORREGIDO: Primero RESUMEN FINANCIERO
        self._add_financial_summary(ws, summary_stats, start_row=7)
        
        # Luego ESTADÍSTICAS del cliente
        self._add_summary_statistics(ws, summary_stats, start_row=18)
        
        # Finalmente REGISTROS FALTANTES (sin color azul)
        self._add_missing_records_summary(ws, summary_stats, start_row=30)
        
        # Gráfico de resumen
        self._add_summary_chart(ws, summary_stats, start_row=35)
        
        # Ajustar columnas
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws.column_dimensions[col].width = 20
    
    def _add_financial_summary(self, ws, summary_stats: Dict, start_row: int):
        """Agrega resumen financiero (CORREGIDO PARA KIOSKO)"""
        
        ws[f'A{start_row}'] = "RESUMEN FINANCIERO"
        ws[f'A{start_row}'].font = self.excel_styles['subtitle']['font']
        ws[f'A{start_row}'].border = Border()
        
        # 🔧 FIX KIOSKO: Datos financieros con manejo mejorado
        currency = self.config['currency_symbol']
        
        # 🔧 CORECCIÓN CRÍTICA: Para KIOSKO usar el campo correcto
        if self.client_type == 'KIOSKO':
            # Para KIOSKO, buscar en múltiples campos posibles
            source_value = 0
            possible_source_fields = [
                'total_valor_kiosko',   # Campo específico KIOSKO
                'total_valor_oxxo',     # Campo de compatibilidad
                'kiosko_total_amount',  # Campo alternativo
                f'total_valor_{self.client_type.lower()}'  # Campo genérico
            ]
            
            for field in possible_source_fields:
                if field in summary_stats and summary_stats[field] != 0:
                    source_value = summary_stats[field]
                    logger.info(f"🔧 KIOSKO Resumen: Usando {field} = ${source_value:,.2f}")
                    break
            
            if source_value == 0:
                logger.error(f"🚨 KIOSKO: No se encontró valor fuente en summary_stats")
                # Buscar cualquier campo que contenga 'total' y 'valor'
                for key, value in summary_stats.items():
                    if ('total' in key.lower() and 'valor' in key.lower() and 
                        isinstance(value, (int, float)) and value != 0):
                        source_value = value
                        logger.info(f"🆘 KIOSKO RESCUE: Usando {key} = ${value:,.2f}")
                        break
            
            source_field_display = f"Total Importe {self.client_type}"
            source_field_value = source_value
        else:
            # Para OXXO usar lógica original
            source_field = f'total_valor_{self.client_type.lower()}'
            source_field_display = f"Total Importe {self.client_type}"
            source_field_value = summary_stats.get(source_field, 0)
        
        # Datos financieros
        financial_data = [
            [source_field_display, source_field_value],
            ["Total Importe Registrado", summary_stats.get('total_valor_looker', 0)],
            ["Diferencia Neta", summary_stats.get('total_diferencia', 0)],
            ["Diferencia Promedio (%)", summary_stats.get('avg_diferencia_porcentaje', 0)],
            ["Máxima Diferencia", summary_stats.get('max_diferencia_abs', 0)]
        ]
        
        # Para KIOSKO, agregar información de devoluciones si existen
        if self.client_type == 'KIOSKO' and summary_stats.get('returns_count', 0) > 0:
            financial_data.insert(-2, ["Total Devoluciones", summary_stats.get('returns_total_amount', 0)])
        
        current_row = start_row + 2
        for description, value in financial_data:
            ws[f'A{current_row}'] = description
            ws[f'B{current_row}'] = value
            ws[f'A{current_row}'].font = Font(bold=True)
            ws[f'A{current_row}'].border = Border()
            ws[f'B{current_row}'].border = Border()
            
            if "%" in description:
                ws[f'B{current_row}'].number_format = '0.00"%"'
                if isinstance(value, (int, float)) and abs(value) < 1:
                    ws[f'B{current_row}'] = value
                elif isinstance(value, (int, float)) and abs(value) > 1:
                    ws[f'B{current_row}'] = value / 100
            else:
                ws[f'B{current_row}'].number_format = f'"{currency}"#,##0.00'
                
                if "Diferencia" in description and isinstance(value, (int, float)):
                    if value > 0:
                        ws[f'B{current_row}'].font = Font(color=self.config['colors']['error'])
                    elif value < 0:
                        ws[f'B{current_row}'].font = Font(color=self.config['colors']['success'])
            
            current_row += 1
    
    def _add_summary_statistics(self, ws, summary_stats: Dict, start_row: int):
        """Agrega estadísticas de resumen con título correcto del cliente"""
        
        ws[f'A{start_row}'] = self.config['stats_title']
        ws[f'A{start_row}'].font = self.excel_styles['subtitle']['font']
        ws[f'A{start_row}'].border = Border()
        
        identifier_plural = self.config['identifier_plural']
        
        # Para KIOSKO, mostrar información de devoluciones si las hay
        total_records = summary_stats.get('total_records', 0)
        returns_count = summary_stats.get('returns_count', 0)
        
        stats_data = [
            [f"Total {identifier_plural} Procesados", total_records, ""],
        ]
        
        # Agregar devoluciones solo para KIOSKO
        if self.client_type == 'KIOSKO' and returns_count > 0:
            stats_data.extend([
                ["Devoluciones (Informativos)", returns_count, "↩️"],
                [f"Total con Devoluciones", total_records + returns_count, ""],
            ])
        
        stats_data.extend([
            ["", "", ""],
            ["CONCILIACIÓN", "", ""],
            [f"{identifier_plural} Conciliados", summary_stats.get('exact_matches', 0), "✅"],
            ["Dentro de Tolerancia", summary_stats.get('within_tolerance', 0), "🟢"],
            ["Diferencias Menores", summary_stats.get('minor_differences', 0), "🟡"],
            ["Diferencias Significativas", summary_stats.get('major_differences', 0), "🔴"],
            ["", "", ""],
            ["TASA DE ÉXITO", f"{summary_stats.get('reconciliation_rate', 0):.1f}%", "🎯"]
        ])
        
        current_row = start_row + 2
        for description, value, icon in stats_data:
            if description and not description.isupper():
                ws[f'A{current_row}'] = description
                ws[f'B{current_row}'] = value
                ws[f'C{current_row}'] = icon
                ws[f'A{current_row}'].font = Font(bold=True)
                ws[f'A{current_row}'].border = Border()
                ws[f'B{current_row}'].border = Border()
                ws[f'C{current_row}'].border = Border()
                
                if isinstance(value, (int, float)) and value != 0:
                    if description == "TASA DE ÉXITO":
                        ws[f'B{current_row}'].number_format = '0.0"%"'
                    else:
                        ws[f'B{current_row}'].number_format = '0'
            
            elif description.isupper():
                ws[f'A{current_row}'] = description
                ws[f'A{current_row}'].font = Font(bold=True, color=self.config['colors']['info'])
                ws[f'A{current_row}'].border = Border()
            
            current_row += 1
    
    def _add_missing_records_summary(self, ws, summary_stats: Dict, start_row: int):
        """Agrega resumen de registros faltantes (SIN COLOR AZUL)"""
        
        ws[f'A{start_row}'] = "REGISTROS FALTANTES"
        ws[f'A{start_row}'].font = Font(bold=True, size=14)
        ws[f'A{start_row}'].border = Border()
        
        # Adaptarse a los nombres de campos específicos de cada cliente
        missing_in_looker = summary_stats.get('missing_in_looker', 0)
        
        if self.client_type == 'KIOSKO':
            missing_in_source = summary_stats.get('missing_in_kiosko', summary_stats.get('missing_in_oxxo', 0))
        else:
            missing_in_source = summary_stats.get(f'missing_in_{self.client_type.lower()}', 0)
        
        missing_data = [
            ["No registrados en sistema", missing_in_looker, "❌"],
            [self.config['missing_source_label'], missing_in_source, "❌"]
        ]
        
        current_row = start_row + 2
        for description, value, icon in missing_data:
            ws[f'A{current_row}'] = description
            ws[f'B{current_row}'] = value
            ws[f'C{current_row}'] = icon
            ws[f'A{current_row}'].font = Font(bold=True)
            ws[f'A{current_row}'].border = Border()
            ws[f'B{current_row}'].border = Border()
            ws[f'C{current_row}'].border = Border()
            
            if isinstance(value, (int, float)):
                ws[f'B{current_row}'].number_format = '0'
            
            current_row += 1
    
    def _create_detailed_results_sheet(self, wb: openpyxl.Workbook, 
                                     reconciliation_results: pd.DataFrame):
        """Crea hoja de resultados detallados CON ENCABEZADOS CORREGIDOS Y VALIDACIÓN"""
        
        ws = wb.create_sheet("Resultados Detallados")
        ws.sheet_view.showGridLines = False
        
        if len(reconciliation_results) == 0:
            ws['A1'] = "No hay datos para mostrar"
            ws['A1'].font = Font(size=14, bold=True, color=self.config['colors']['error'])
            ws['A1'].border = Border()
            return
        
        # DEBUGGING: Log estructura del DataFrame
        logger.info(f"🔍 DEBUG: DataFrame shape: {reconciliation_results.shape}")
        logger.info(f"🔍 DEBUG: DataFrame columns: {list(reconciliation_results.columns)}")
        logger.info(f"🔍 DEBUG: DataFrame dtypes: {reconciliation_results.dtypes.to_dict()}")
        
        # Preparar datos para mostrar CON VALIDACIÓN MEJORADA
        try:
            df_display = self._prepare_display_data_safe(reconciliation_results.copy())
        except Exception as e:
            logger.error(f"❌ Error preparando datos: {e}")
            # Fallback: usar datos originales con columnas básicas
            df_display = self._prepare_fallback_display_data(reconciliation_results.copy())
        
        # Título
        ws['A1'] = f"RESULTADOS DETALLADOS - {self.config['display_name']}"
        ws['A1'].font = self.excel_styles['title']['font']
        ws['A1'].border = Border()
        ws.merge_cells(f'A1:{chr(65 + len(df_display.columns) - 1)}1')
        
        # Headers en la fila 2
        header_row = 2
        for col_idx, column_name in enumerate(df_display.columns, 1):
            cell = ws.cell(row=header_row, column=col_idx, value=column_name)
            cell.font = self.excel_styles['header']['font']
            cell.fill = self.excel_styles['header']['fill']
            cell.alignment = self.excel_styles['header']['alignment']
            cell.border = Border()
        
        # Datos empezando en la fila 3
        data_start_row = 3
        for row_idx, (_, row_data) in enumerate(df_display.iterrows(), data_start_row):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = Border()
        
        # Aplicar formatos
        self._apply_detailed_formatting(ws, df_display, start_row=data_start_row)
        self._auto_adjust_columns(ws)
    
    def _prepare_display_data_safe(self, df: pd.DataFrame) -> pd.DataFrame:
        """🔧 FIX KIOSKO: Prepara datos para mostrar CON VALIDACIÓN ROBUSTA"""
        
        logger.info(f"🔧 Preparando datos para display (SAFE MODE con FIX KIOSKO)")
        
        if len(df) == 0:
            logger.warning("⚠️ DataFrame vacío")
            return pd.DataFrame()
        
        # Mapeo de categorías
        category_mapping = self.config['categories']
        
        # Mapeo de columnas con VALIDACIÓN
        base_columns_mapping = {
            'id_matching': self.config['identifier_field'],
            'diferencia_valor': 'Diferencia',
            'diferencia_porcentaje': '% Diferencia',
            'categoria': 'Estatus',
            'match_confidence': 'Confianza %'
        }
        
        # 🔧 FIX PRINCIPAL: Mapeo específico para campos de valor según cliente
        if self.client_type == 'KIOSKO':
            # ORDEN DE PRIORIDAD para encontrar valores KIOSKO
            value_columns_mapping = {
                'valor_source_clean': self.config['value_field_source'],  # PRINCIPAL
                'total_venta_kiosko': self.config['value_field_source'],  # BACKUP
                'valor_oxxo_clean': self.config['value_field_source'],    # COMPATIBILIDAD
                'valor_kiosko_clean': self.config['value_field_source'],  # IDEAL
                'valor_looker_clean': self.config['value_field_looker'],
            }
            logger.info(f"🔧 KIOSKO: Buscando valores en orden de prioridad...")
        else:
            # Para OXXO mantener lógica original
            value_columns_mapping = {
                'valor_source_clean': self.config['value_field_source'],
                'valor_oxxo_clean': self.config['value_field_source'],
                'valor_looker_clean': self.config['value_field_looker'],
            }
        
        # Combinar mapeos
        columns_mapping = {**base_columns_mapping, **value_columns_mapping}
        
        # PASO 1: Encontrar columnas disponibles con VALIDACIÓN MEJORADA
        available_columns = {}
        source_value_found = False
        
        for original_name, display_name in columns_mapping.items():
            if original_name in df.columns and not df[original_name].empty:
                available_columns[original_name] = display_name
                logger.info(f"   ✅ Columna encontrada: {original_name} -> {display_name}")
                
                # Marcar si encontramos valor fuente para KIOSKO
                if (self.client_type == 'KIOSKO' and 
                    original_name in ['valor_source_clean', 'total_venta_kiosko', 'valor_oxxo_clean', 'valor_kiosko_clean']):
                    source_value_found = True
                    logger.info(f"   🎯 KIOSKO: Valor fuente encontrado en {original_name}")
        
        # 🚨 VERIFICACIÓN CRÍTICA PARA KIOSKO
        if self.client_type == 'KIOSKO' and not source_value_found:
            logger.error(f"🚨 KIOSKO ERROR: No se encontró columna de valor fuente")
            # 🆘 RESCUE: Buscar cualquier columna con valores KIOSKO
            for col in df.columns:
                if ('kiosko' in col.lower() or 'source' in col.lower()) and ('total' in col.lower() or 'valor' in col.lower()):
                    if not df[col].empty and df[col].sum() != 0:
                        available_columns[col] = self.config['value_field_source']
                        logger.info(f"   🆘 RESCUE: Usando columna {col} como valor KIOSKO")
                        source_value_found = True
                        break
        
        if not available_columns:
            logger.error("❌ No se encontraron columnas válidas")
            return self._prepare_fallback_display_data(df)
        
        # PASO 2: Seleccionar y renombrar columnas
        try:
            df_display = df[list(available_columns.keys())].copy()
            df_display = df_display.rename(columns=available_columns)
            logger.info(f"   📊 DataFrame display creado: {df_display.shape}")
            
            # 🔍 VERIFICACIÓN FINAL PARA KIOSKO
            if self.client_type == 'KIOSKO':
                kiosko_value_col = self.config['value_field_source']
                if kiosko_value_col in df_display.columns:
                    total_kiosko = df_display[kiosko_value_col].sum()
                    logger.info(f"   💰 KIOSKO: Total en {kiosko_value_col}: ${total_kiosko:,.2f}")
                
        except Exception as e:
            logger.error(f"❌ Error seleccionando columnas: {e}")
            return self._prepare_fallback_display_data(df)
        
        # PASO 3: Traducir categorías
        if 'Estatus' in df_display.columns:
            df_display['Estatus'] = df_display['Estatus'].map(category_mapping).fillna(df_display['Estatus'])
        
        # PASO 4: Convertir tipos de datos
        for column in df_display.columns:
            try:
                if self.config['identifier_field'] in column:
                    df_display[column] = pd.to_numeric(df_display[column], errors='coerce').fillna(0).astype(int)
                elif any(keyword in column for keyword in ['Importe', 'Diferencia']) and '%' not in column:
                    df_display[column] = pd.to_numeric(df_display[column], errors='coerce').fillna(0)
                elif '% Diferencia' in column:
                    numeric_values = pd.to_numeric(df_display[column], errors='coerce').fillna(0)
                    df_display[column] = numeric_values / 100
                elif 'Confianza %' in column:
                    numeric_values = pd.to_numeric(df_display[column], errors='coerce').fillna(0)
                    df_display[column] = numeric_values / 100
            except Exception as e:
                logger.error(f"❌ Error procesando columna {column}: {e}")
        
        logger.info(f"✅ Datos preparados exitosamente: {df_display.shape}")
        return df_display
    def _prepare_fallback_display_data(self, df: pd.DataFrame) -> pd.DataFrame:
        """🆘 FALLBACK: Prepara datos básicos cuando falla el procesamiento principal"""
        
        logger.warning("🆘 Usando modo FALLBACK para preparar datos")
        
        # Seleccionar columnas básicas que deberían existir
        basic_columns = []
        
        # Buscar columna ID
        id_candidates = ['id_matching', 'identificador_unico', 'pedido_adicional', 'ticket']
        for candidate in id_candidates:
            if candidate in df.columns:
                basic_columns.append((candidate, self.config['identifier_field']))
                break
        
        # Buscar columna de categoría
        if 'categoria' in df.columns:
            basic_columns.append(('categoria', 'Estatus'))
        elif 'match_type' in df.columns:
            basic_columns.append(('match_type', 'Tipo'))
        
        # Buscar columnas de valor
        value_candidates = ['valor_oxxo_clean', 'valor_kiosko_clean', 'valor_source_clean', 'total_venta']
        for candidate in value_candidates:
            if candidate in df.columns:
                basic_columns.append((candidate, self.config['value_field_source']))
                break
        
        # Buscar columna de diferencia
        if 'diferencia_valor' in df.columns:
            basic_columns.append(('diferencia_valor', 'Diferencia'))
        
        if not basic_columns:
            # Ultra fallback: usar las primeras 5 columnas
            logger.warning("🚨 Ultra fallback: usando primeras columnas")
            return df.iloc[:, :min(5, len(df.columns))].copy()
        
        # Crear DataFrame con columnas básicas
        df_basic = pd.DataFrame()
        for original_col, display_name in basic_columns:
            if original_col in df.columns:
                df_basic[display_name] = df[original_col]
        
        logger.info(f"🆘 Fallback completado: {df_basic.shape}")
        return df_basic
    
    def _apply_detailed_formatting(self, ws, df_display: pd.DataFrame, start_row: int = 3):
        """Aplica formato específico CON PORCENTAJES CORREGIDOS"""
        
        column_formats = {
            self.config['identifier_field']: '0',
            self.config['value_field_source']: self.excel_styles['currency']['number_format'],
            self.config['value_field_looker']: self.excel_styles['currency']['number_format'],
            'Diferencia': self.excel_styles['currency']['number_format'],
            '% Diferencia': '0.00%',
            'Confianza %': '0"%"',
            'Estatus': '@'
        }
        
        for col_idx, column_name in enumerate(df_display.columns, 1):
            if column_name in column_formats:
                format_code = column_formats[column_name]
                
                for row_idx in range(start_row, len(df_display) + start_row):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.number_format = format_code
                    cell.border = Border()
                    
                    if column_name == 'Estatus':
                        self._apply_status_colors(cell)
                    elif column_name == 'Diferencia':
                        self._apply_difference_colors(cell)
                    elif column_name == '% Diferencia':
                        self._apply_percentage_colors(cell)
    
    def _apply_status_colors(self, cell):
        """Aplica colores según el estatus"""
        
        status_colors = {
            f'{self.config["identifier_name"]} Conciliado': 'C6EFCE',
            'Ticket Conciliado': 'C6EFCE',
            'Pedido Conciliado': 'C6EFCE',
            'Dentro de Tolerancia': 'D9EAD3',
            'Diferencia Menor': 'F2F2F2',
            'Diferencia Mayor': 'FFEB9C',
            f'{self.config["identifier_name"]} no registrado': 'FFC7CE',
            'Ticket no registrado': 'FFC7CE',
            'Pedido no registrado': 'FFC7CE',
            f'Faltante en {self.client_type}': 'FFC7CE',
            'Devolución (Informativo)': 'E1F5FE',  # NUEVO: color para devoluciones
        }
        
        if cell.value in status_colors:
            cell.fill = PatternFill(start_color=status_colors[cell.value], 
                                  end_color=status_colors[cell.value], 
                                  fill_type="solid")
            cell.border = Border()
    
    def _apply_difference_colors(self, cell):
        """Aplica colores según el valor de la diferencia"""
        try:
            value = float(cell.value) if cell.value else 0
            
            if value == 0:
                cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type="solid")
            elif abs(value) > 100:
                cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type="solid")
            elif abs(value) > 50:
                cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type="solid")
            
            cell.border = Border()
        except (ValueError, TypeError):
            pass
    
    def _apply_percentage_colors(self, cell):
        """Aplica colores según el porcentaje de diferencia"""
        try:
            value = float(cell.value) if cell.value else 0
            
            if abs(value) == 0:
                cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type="solid")
            elif abs(value) > 0.1:
                cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type="solid")
            elif abs(value) > 0.05:
                cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type="solid")
            
            cell.border = Border()
        except (ValueError, TypeError):
            pass
    
    def _create_analysis_sheet(self, wb: openpyxl.Workbook, 
                             reconciliation_results: pd.DataFrame, summary_stats: Dict):
        """Crea hoja de análisis avanzado - RESTAURADO COMPLETO"""
        
        ws = wb.create_sheet("Análisis")
        ws.sheet_view.showGridLines = False
        
        ws['A1'] = f"ANÁLISIS AVANZADO - {self.config['display_name']}"
        ws['A1'].font = self.excel_styles['title']['font']
        ws['A1'].border = Border()
        
        # Análisis por rangos de diferencias - RESTAURADO
        self._add_difference_analysis(ws, reconciliation_results, start_row=4)
        
        # Análisis de tendencias - RESTAURADO
        if len(reconciliation_results) > 10:
            self._add_trends_analysis(ws, reconciliation_results, start_row=15)
        
        # Recomendaciones - RESTAURADO
        self._add_recommendations(ws, summary_stats, start_row=25)
    
    def _add_difference_analysis(self, ws, df: pd.DataFrame, start_row: int):
        """Agrega análisis de diferencias por rangos - RESTAURADO"""
        
        ws[f'A{start_row}'] = "ANÁLISIS DE DIFERENCIAS"
        ws[f'A{start_row}'].font = self.excel_styles['subtitle']['font']
        ws[f'A{start_row}'].border = Border()
        
        if 'diferencia_absoluta' in df.columns:
            # Definir rangos de diferencias
            ranges = [
                (0, 0, "Sin diferencia"),
                (0.01, 25, "Diferencias pequeñas (≤$25)"),
                (25.01, 100, "Diferencias medianas ($25-$100)"),
                (100.01, 500, "Diferencias grandes ($100-$500)"),
                (500.01, float('inf'), "Diferencias muy grandes (>$500)")
            ]
            
            # Analizar rangos
            analysis_data = [["Rango", "Cantidad", "% del Total"]]
            total_records = len(df)
            
            for min_val, max_val, description in ranges:
                if max_val == float('inf'):
                    count = len(df[df['diferencia_absoluta'] > min_val])
                else:
                    count = len(df[(df['diferencia_absoluta'] >= min_val) & (df['diferencia_absoluta'] <= max_val)])
                
                percentage = (count / total_records * 100) if total_records > 0 else 0
                analysis_data.append([description, count, f"{percentage:.1f}%"])
            
            # Agregar datos
            current_row = start_row + 2
            for row_data in analysis_data:
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=current_row, column=col_idx, value=value)
                    cell.border = Border()
                    if current_row == start_row + 2:  # Header
                        cell.font = Font(bold=True)
                current_row += 1
        else:
            ws[f'A{start_row + 2}'] = "No hay datos de diferencias para analizar"
            ws[f'A{start_row + 2}'].border = Border()
    
    def _add_trends_analysis(self, ws, df: pd.DataFrame, start_row: int):
        """Agrega análisis de tendencias - RESTAURADO"""
        
        ws[f'A{start_row}'] = "ANÁLISIS DE PATRONES"
        ws[f'A{start_row}'].font = self.excel_styles['subtitle']['font']
        ws[f'A{start_row}'].border = Border()
        
        # Análisis de patrones de conciliación
        if 'categoria' in df.columns:
            category_analysis = df['categoria'].value_counts()
            
            current_row = start_row + 2
            ws[f'A{current_row}'] = "Distribución de Resultados:"
            ws[f'A{current_row}'].font = Font(bold=True)
            ws[f'A{current_row}'].border = Border()
            
            current_row += 1
            for category, count in category_analysis.items():
                percentage = (count / len(df)) * 100
                translated_category = self.config['categories'].get(category, category)
                ws[f'A{current_row}'] = f"• {translated_category}: {count} ({percentage:.1f}%)"
                ws[f'A{current_row}'].border = Border()
                current_row += 1
        
        # Análisis de valores
        if 'diferencia_absoluta' in df.columns:
            current_row += 1
            ws[f'A{current_row}'] = "Estadísticas de Diferencias:"
            ws[f'A{current_row}'].font = Font(bold=True)
            ws[f'A{current_row}'].border = Border()
            
            stats_data = [
                ["Promedio", df['diferencia_absoluta'].mean()],
                ["Mediana", df['diferencia_absoluta'].median()],
                ["Máximo", df['diferencia_absoluta'].max()],
                ["Mínimo", df['diferencia_absoluta'].min()]
            ]
            
            for stat_name, stat_value in stats_data:
                current_row += 1
                ws[f'A{current_row}'] = f"• {stat_name}: ${stat_value:.2f}"
                ws[f'A{current_row}'].border = Border()
    
    def _add_recommendations(self, ws, summary_stats: Dict, start_row: int):
        """Agrega recomendaciones inteligentes - RESTAURADO CON MEJORAS PARA KIOSKO"""
        
        ws[f'A{start_row}'] = "RECOMENDACIONES"
        ws[f'A{start_row}'].font = self.excel_styles['subtitle']['font']
        ws[f'A{start_row}'].border = Border()
        
        # Generar recomendaciones inteligentes
        recommendations = []
        
        reconciliation_rate = summary_stats.get('reconciliation_rate', 0)
        major_differences = summary_stats.get('major_differences', 0)
        
        # Adaptarse a los nombres específicos de cada cliente
        if self.client_type == 'KIOSKO':
            missing_in_source = summary_stats.get('missing_in_kiosko', summary_stats.get('missing_in_oxxo', 0))
        else:
            missing_in_source = summary_stats.get(f'missing_in_{self.client_type.lower()}', 0)
            
        missing_in_looker = summary_stats.get('missing_in_looker', 0)
        missing_records = missing_in_looker + missing_in_source
        
        exact_matches = summary_stats.get('exact_matches', 0)
        total_records = summary_stats.get('total_records', 0)
        
        # Recomendaciones específicas para KIOSKO con devoluciones
        if self.client_type == 'KIOSKO':
            returns_count = summary_stats.get('returns_count', 0)
            if returns_count > 0:
                returns_amount = summary_stats.get('returns_total_amount', 0)
                recommendations.append(f"↩️ Se procesaron {returns_count} devoluciones por ${abs(returns_amount):,.2f} como informativos.")
                recommendations.append("📝 Las devoluciones no afectan la conciliación principal pero se incluyen en el reporte.")
        
        # Recomendaciones basadas en tasa de conciliación
        if reconciliation_rate >= 95:
            recommendations.append("✅ Excelente tasa de conciliación. El proceso está funcionando correctamente.")
        elif reconciliation_rate >= 85:
            recommendations.append("🟡 Buena tasa de conciliación, pero hay margen de mejora.")
            recommendations.append("💡 Revisar procesos de captura de datos para aumentar precisión.")
        elif reconciliation_rate >= 70:
            recommendations.append("🟠 Tasa de conciliación moderada. Se requiere atención.")
            recommendations.append("🔍 Investigar causas de discrepancias en el proceso.")
        else:
            recommendations.append("🔴 Tasa de conciliación baja. Se requiere revisión urgente del proceso.")
            recommendations.append("⚠️ Posibles problemas en la captura o transmisión de datos.")
        
        # Recomendaciones basadas en diferencias
        if major_differences > 0:
            percentage = (major_differences / total_records * 100) if total_records > 0 else 0
            recommendations.append(f"⚠️ Revisar {major_differences} {self.config['identifier_plural'].lower()} con diferencias significativas ({percentage:.1f}% del total).")
            
            if percentage > 10:
                recommendations.append("🚨 Alto porcentaje de diferencias significativas requiere investigación inmediata.")
        
        # Recomendaciones basadas en registros faltantes
        if missing_records > 0:
            missing_percentage = (missing_records / total_records * 100) if total_records > 0 else 0
            recommendations.append(f"📋 Investigar {missing_records} registros faltantes ({missing_percentage:.1f}% del total).")
            
            if missing_in_looker > missing_in_source:
                recommendations.append("📤 Mayor cantidad de registros no enviados al sistema. Revisar proceso de transmisión.")
            elif missing_in_source > missing_in_looker:
                recommendations.append("📥 Registros en sistema sin correspondencia en origen. Verificar filtros de datos.")
        
        # Recomendaciones basadas en estado ideal
        if missing_records == 0 and major_differences == 0:
            recommendations.append("🎉 Conciliación perfecta. Todos los registros están balanceados.")
            recommendations.append("✨ Mantener los procesos actuales para conservar esta calidad.")
        
        # Recomendaciones específicas por cliente
        if self.client_type == 'OXXO' and exact_matches > 0:
            recommendations.append("📦 Para OXXO: Verificar que la agrupación de productos múltiples sea correcta.")
        elif self.client_type == 'KIOSKO' and exact_matches > 0:
            recommendations.append("🧊 Para KIOSKO: Validar normalización de IDs de tickets con guiones.")
            if summary_stats.get('returns_count', 0) > 0:
                recommendations.append("↩️ Para KIOSKO: Las devoluciones se manejan por separado y no requieren conciliación.")
        
        # Agregar recomendaciones al worksheet
        current_row = start_row + 2
        for rec in recommendations:
            ws[f'A{current_row}'] = rec
            ws[f'A{current_row}'].border = Border()
            # Ajustar ancho para recomendaciones largas
            if len(rec) > 80:
                ws.row_dimensions[current_row].height = 30
            current_row += 1
    
    def _create_differences_sheet(self, wb: openpyxl.Workbook, 
                                reconciliation_results: pd.DataFrame):
        """Crea hoja específica para diferencias encontradas"""
        
        ws = wb.create_sheet("Diferencias")
        ws.sheet_view.showGridLines = False
        
        if 'categoria' in reconciliation_results.columns:
            differences = reconciliation_results[
                reconciliation_results['categoria'].isin(['MINOR_DIFFERENCE', 'MAJOR_DIFFERENCE'])
            ]
        else:
            differences = pd.DataFrame()
        
        if len(differences) == 0:
            ws['A1'] = "✅ No se encontraron diferencias significativas"
            ws['A1'].font = Font(size=14, bold=True, color=self.config['colors']['success'])
            ws['A1'].border = Border()
            return
        
        ws['A1'] = f"DIFERENCIAS ENCONTRADAS - {len(differences)} {self.config['identifier_plural'].lower()}"
        ws['A1'].font = Font(size=14, bold=True, color=self.config['colors']['warning'])
        ws['A1'].border = Border()
        
        # Preparar y agregar datos (con validación)
        try:
            df_display = self._prepare_display_data_safe(differences.copy())
        except:
            df_display = self._prepare_fallback_display_data(differences.copy())
        
        # Headers en fila 2
        for col_idx, column_name in enumerate(df_display.columns, 1):
            cell = ws.cell(row=2, column=col_idx, value=column_name)
            cell.font = self.excel_styles['header']['font']
            cell.fill = self.excel_styles['header']['fill']
            cell.alignment = self.excel_styles['header']['alignment']
            cell.border = Border()
        
        # Datos empezando en fila 3
        for row_idx, (_, row_data) in enumerate(df_display.iterrows(), 3):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = Border()
        
        self._apply_detailed_formatting(ws, df_display, start_row=3)
        self._auto_adjust_columns(ws)
    
    def _create_missing_records_sheet(self, wb: openpyxl.Workbook, 
                                    reconciliation_results: pd.DataFrame):
        """Crea hoja para registros faltantes"""
        
        ws = wb.create_sheet("Registros Faltantes")
        ws.sheet_view.showGridLines = False
        
        if 'categoria' in reconciliation_results.columns:
            missing = reconciliation_results[
                reconciliation_results['categoria'].str.contains('MISSING', na=False)
            ]
        else:
            missing = pd.DataFrame()
        
        if len(missing) == 0:
            ws['A1'] = "✅ No se encontraron registros faltantes"
            ws['A1'].font = Font(size=14, bold=True, color=self.config['colors']['success'])
            ws['A1'].border = Border()
            return
        
        ws['A1'] = f"REGISTROS FALTANTES - {len(missing)} {self.config['identifier_plural'].lower()}"
        ws['A1'].font = Font(size=14, bold=True, color=self.config['colors']['error'])
        ws['A1'].border = Border()
        
        # Preparar y agregar datos (con validación)
        try:
            df_display = self._prepare_display_data_safe(missing.copy())
        except:
            df_display = self._prepare_fallback_display_data(missing.copy())
        
        # Headers en fila 2
        for col_idx, column_name in enumerate(df_display.columns, 1):
            cell = ws.cell(row=2, column=col_idx, value=column_name)
            cell.font = self.excel_styles['header']['font']
            cell.fill = self.excel_styles['header']['fill']
            cell.alignment = self.excel_styles['header']['alignment']
            cell.border = Border()
        
        # Datos empezando en fila 3
        for row_idx, (_, row_data) in enumerate(df_display.iterrows(), 3):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = Border()
        
        self._apply_detailed_formatting(ws, df_display, start_row=3)
        self._auto_adjust_columns(ws)
    
    def _create_billing_sheet(self, wb: openpyxl.Workbook, 
                            reconciliation_results: pd.DataFrame):
        """Crea hoja con registros listos para facturación"""
        
        ws = wb.create_sheet("Listos para Facturación")
        ws.sheet_view.showGridLines = False
        
        if 'categoria' in reconciliation_results.columns:
            # Para KIOSKO, excluir devoluciones de facturación
            billing_categories = ['EXACT_MATCH', 'WITHIN_TOLERANCE']
            billing_ready = reconciliation_results[
                reconciliation_results['categoria'].isin(billing_categories)
            ]
            
            # Para KIOSKO: filtrar devoluciones adicionales
            if self.client_type == 'KIOSKO' and 'is_return' in reconciliation_results.columns:
                billing_ready = billing_ready[
                    (billing_ready.get('is_return', False) != True)
                ]
        else:
            billing_ready = pd.DataFrame()
        
        if len(billing_ready) == 0:
            ws['A1'] = "❌ No hay registros listos para facturación"
            ws['A1'].font = Font(size=14, bold=True, color=self.config['colors']['error'])
            ws['A1'].border = Border()
            return
        
        # Calcular total para facturación
        value_field = 'valor_oxxo_clean' if self.client_type == 'KIOSKO' else 'valor_source_clean'
        total_amount = billing_ready.get(value_field, pd.Series([0])).sum()
        
        ws['A1'] = f"✅ APROBADOS PARA FACTURACIÓN - {len(billing_ready)} {self.config['identifier_plural'].lower()}"
        ws['A1'].font = Font(size=14, bold=True, color=self.config['colors']['success'])
        ws['A1'].border = Border()
        
        ws['A2'] = f"Importe total: {self.config['currency_symbol']}{total_amount:,.2f}"
        ws['A2'].font = Font(size=12, bold=True)
        ws['A2'].border = Border()
        
        # Nota especial para KIOSKO sobre devoluciones
        if self.client_type == 'KIOSKO':
            ws['A3'] = "📝 Nota: Las devoluciones se excluyen automáticamente de la facturación"
            ws['A3'].font = Font(size=10, italic=True)
            ws['A3'].border = Border()
            start_row = 5
        else:
            start_row = 4
        
        # Preparar y agregar datos (con validación)
        try:
            df_display = self._prepare_display_data_safe(billing_ready.copy())
        except:
            df_display = self._prepare_fallback_display_data(billing_ready.copy())
        
        # Headers
        for col_idx, column_name in enumerate(df_display.columns, 1):
            cell = ws.cell(row=start_row, column=col_idx, value=column_name)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=self.config['colors']['success'], 
                                  end_color=self.config['colors']['success'], 
                                  fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border()
        
        # Datos
        for row_idx, (_, row_data) in enumerate(df_display.iterrows(), start_row + 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = Border()
        
        self._apply_detailed_formatting(ws, df_display, start_row=start_row + 1)
        self._auto_adjust_columns(ws)
    
    def _create_products_analysis_sheet(self, wb: openpyxl.Workbook, 
                                      reconciliation_results: pd.DataFrame):
        """Crea hoja específica para análisis de productos (solo KIOSKO)"""
        
        if self.client_type != 'KIOSKO':
            return
        
        ws = wb.create_sheet("Análisis de Productos")
        ws.sheet_view.showGridLines = False
        
        ws['A1'] = "ANÁLISIS DE PRODUCTOS KIOSKO"
        ws['A1'].font = self.excel_styles['title']['font']
        ws['A1'].border = Border()
        
        # Análisis de devoluciones si existen
        returns_data = reconciliation_results[
            reconciliation_results.get('categoria', '') == 'RETURN_INFORMATIVE'
        ]
        
        if len(returns_data) > 0:
            ws['A3'] = f"DEVOLUCIONES PROCESADAS: {len(returns_data)}"
            ws['A3'].font = Font(bold=True, size=12)
            ws['A3'].border = Border()
            
            total_returns = returns_data.get('valor_oxxo_clean', pd.Series([0])).sum()
            ws['A4'] = f"Importe total devoluciones: ${abs(total_returns):,.2f}"
            ws['A4'].border = Border()
            
            ws['A6'] = "Las devoluciones se procesan automáticamente como:"
            ws['A6'].border = Border()
            ws['A7'] = "• Registros informativos (no afectan conciliación)"
            ws['A7'].border = Border()
            ws['A8'] = "• Se excluyen de facturación automáticamente"
            ws['A8'].border = Border()
            ws['A9'] = "• Se identifican por prefijos (36_, 75_, 99_) o valores negativos"
            ws['A9'].border = Border()
        else:
            ws['A3'] = "✅ No se encontraron devoluciones en este período"
            ws['A3'].font = Font(bold=True, color=self.config['colors']['success'])
            ws['A3'].border = Border()
        
        ws['A11'] = "Análisis detallado por productos de hielo"
        ws['A11'].border = Border()
        ws['A12'] = "Esta sección se puede expandir con análisis específicos de:"
        ws['A12'].border = Border()
        ws['A13'] = "• Productos más vendidos"
        ws['A13'].border = Border()
        ws['A14'] = "• Análisis por tipo de hielo"
        ws['A14'].border = Border()
        ws['A15'] = "• Tendencias de venta"
        ws['A15'].border = Border()
        
        for row in [11, 12, 13, 14, 15]:
            ws[f'A{row}'].font = Font(size=11)
    
    def _auto_adjust_columns(self, ws):
        """Ajusta automáticamente el ancho de las columnas - CORREGIDO para celdas fusionadas"""
        
        max_column = ws.max_column
        
        for col_idx in range(1, max_column + 1):
            max_length = 0
            column_letter = openpyxl.utils.get_column_letter(col_idx)
            
            for row_idx in range(1, ws.max_row + 1):
                try:
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if hasattr(cell, 'value') and cell.value is not None:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            adjusted_width = max(adjusted_width, 10)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _add_summary_chart(self, ws, summary_stats: Dict, start_row: int):
        """Agrega gráfico de resumen de conciliación"""
        
        ws[f'A{start_row}'] = "DISTRIBUCIÓN DE RESULTADOS"
        ws[f'A{start_row}'].font = self.excel_styles['subtitle']['font']
        ws[f'A{start_row}'].border = Border()
        
        # Datos para el gráfico adaptados por cliente
        missing_in_looker = summary_stats.get('missing_in_looker', 0)
        if self.client_type == 'KIOSKO':
            missing_in_source = summary_stats.get('missing_in_kiosko', summary_stats.get('missing_in_oxxo', 0))
        else:
            missing_in_source = summary_stats.get(f'missing_in_{self.client_type.lower()}', 0)
        
        chart_data = [
            ["Categoría", "Cantidad"],
            ["Conciliados", summary_stats.get('exact_matches', 0)],
            ["Tolerancia", summary_stats.get('within_tolerance', 0)],
            ["Dif. Menores", summary_stats.get('minor_differences', 0)],
            ["Dif. Mayores", summary_stats.get('major_differences', 0)],
            ["Faltantes", missing_in_looker + missing_in_source]
        ]
        
        chart_start_row = start_row + 2
        for row_idx, (category, value) in enumerate(chart_data):
            ws[f'E{chart_start_row + row_idx}'] = category
            ws[f'F{chart_start_row + row_idx}'] = value
            ws[f'E{chart_start_row + row_idx}'].border = Border()
            ws[f'F{chart_start_row + row_idx}'].border = Border()
        
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = f"Resultados de Conciliación {self.client_type}"
        chart.y_axis.title = f'Número de {self.config["identifier_plural"]}'
        chart.x_axis.title = 'Categorías'
        
        data = Reference(ws, min_col=6, min_row=chart_start_row + 1, max_row=chart_start_row + 5, max_col=6)
        cats = Reference(ws, min_col=5, min_row=chart_start_row + 1, max_row=chart_start_row + 5, max_col=5)
        chart.add_data(data, titles_from_data=False)
        chart.set_categories(cats)
        
        ws.add_chart(chart, f"A{start_row + 10}")
    
    def generate_csv_reports(self, reconciliation_results: pd.DataFrame, 
                           timestamp: str = None) -> Dict[str, str]:
        """Genera reportes en formato CSV personalizados por cliente"""
        if timestamp is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        csv_files = {}
        
        try:
            df_display = self._prepare_display_data_safe(reconciliation_results.copy())
        except:
            df_display = self._prepare_fallback_display_data(reconciliation_results.copy())
        
        complete_csv = self.output_dir / f"conciliacion_{self.client_type.lower()}_{timestamp}.csv"
        df_display.to_csv(complete_csv, index=False, encoding='utf-8-sig')
        csv_files['completo'] = str(complete_csv)
        
        # Reportes específicos...
        approved_statuses = [self.config['categories'].get(cat, cat) for cat in ['EXACT_MATCH', 'WITHIN_TOLERANCE']]
        if 'Estatus' in df_display.columns:
            billing_ready = df_display[df_display['Estatus'].isin(approved_statuses)]
            
            if len(billing_ready) > 0:
                billing_csv = self.output_dir / f"facturacion_{self.client_type.lower()}_{timestamp}.csv"
                billing_ready.to_csv(billing_csv, index=False, encoding='utf-8-sig')
                csv_files['facturacion'] = str(billing_csv)
        
        difference_statuses = [self.config['categories'].get(cat, cat) for cat in ['MINOR_DIFFERENCE', 'MAJOR_DIFFERENCE']]
        if 'Estatus' in df_display.columns:
            differences = df_display[df_display['Estatus'].isin(difference_statuses)]
            
            if len(differences) > 0:
                diff_csv = self.output_dir / f"diferencias_{self.client_type.lower()}_{timestamp}.csv"
                differences.to_csv(diff_csv, index=False, encoding='utf-8-sig')
                csv_files['diferencias'] = str(diff_csv)
        
        logger.info(f"✅ Reportes CSV generados para {self.client_type}: {list(csv_files.keys())}")
        return csv_files
    
    def generate_summary_json(self, summary_stats: Dict, timestamp: str = None) -> str:
        """Genera resumen en formato JSON para integración con otros sistemas"""
        if timestamp is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        json_data = {
            'client_type': self.client_type,
            'report_timestamp': datetime.now().isoformat(),
            'summary': summary_stats,
            'client_config': self.config,
            'file_timestamp': timestamp
        }
        
        json_file = self.output_dir / f"resumen_{self.client_type.lower()}_{timestamp}.json"
        with open(json_file, 'w', encoding='utf-8') as f:
            import json
            json.dump(json_data, f, indent=2, ensure_ascii=False, default=str)
        
        logger.info(f"📄 Resumen JSON generado: {json_file}")
        return str(json_file)


# Función de conveniencia para crear reportes
def create_client_report(client_type: str, reconciliation_results: pd.DataFrame, 
                        summary_stats: Dict, timestamp: str = None) -> Dict[str, str]:
    """Función de conveniencia para generar todos los reportes de un cliente"""
    
    generator = ReportGenerator(client_type)
    
    files_generated = {}
    
    excel_file = generator.generate_complete_report(reconciliation_results, summary_stats, timestamp)
    files_generated['excel'] = excel_file
    
    csv_files = generator.generate_csv_reports(reconciliation_results, timestamp)
    files_generated.update(csv_files)
    
    json_file = generator.generate_summary_json(summary_stats, timestamp)
    files_generated['json'] = json_file
    
    return files_generated


if __name__ == "__main__":
    print("📊 TESTING REPORT GENERATOR CORREGIDO - VERSIÓN CON FIX")
    print("=" * 60)
    
    for client in ['OXXO', 'KIOSKO']:
        print(f"\n🧪 Testing {client}...")
        try:
            generator = ReportGenerator(client)
            print(f"✅ {client}: ReportGenerator creado exitosamente")
            print(f"   Configuración: {generator.config['display_name']}")
            print(f"   Identificador: {generator.config['identifier_name']}")
            print(f"   Título estadísticas: {generator.config['stats_title']}")
        except Exception as e:
            print(f"❌ {client}: Error - {e}")
    
    print("\n🎉 Testing completado - VERSION CON FIX PARA TypeError")